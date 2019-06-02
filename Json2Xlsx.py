import xlrd
import json
import os
from openpyxl import Workbook
import time

wb = Workbook()
ws = wb.active

cols = []


def json2xlsx(fileName):
        wb1 = Workbook()
        sheet = wb1.active
        data = json.load(open(fileName + ".json", "r", encoding="utf-8"))
        values = list(data.values())
        for c, i in enumerate(values[0].keys()):
            sheet.cell(row=1, column=c + 1, value=i)
        for r, i in enumerate(values):
            row = r + 2
            for c, d in enumerate(values[r]):
                sheet.cell(row=row, column=c + 1, value=values[r].get(d))
        timeStrap = time.strftime("%a%b%d_%H_%M_%S_%Y", time.localtime());  # time strap
        fileOutName = "test_" + timeStrap + ".xlsx";
        wb1.save(fileOutName)


if __name__ == '__main__':
    jsfile = "test"
    excfile = "testOut.xlsx"
    json2xlsx(jsfile)
