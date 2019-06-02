# -*- coding: utf-8 -*-
from flask import Flask, abort, request, jsonify
from flask_cors import CORS, cross_origin
from openpyxl import Workbook
import json
import xlrd
import os
import time

app = Flask(__name__)
CORS(app)

# 数据预存放,模拟数据库DB
tasks = [
    {
        'id': 1,
        'info': 'this is info 1'
    },
    {
        'id': 2,
        'info': 'this is info 2'
    }
]


@app.route('/index')
@cross_origin(supports_credentials=True)
def index():
    return "This is the front page"


@app.route('/add_task/', methods=['POST'])
@cross_origin(supports_credentials=True)
def add_task():
    if not request.json or 'id' not in request.json or 'info' not in request.json:
        abort(400)
    task = {
        'id': request.json['id'],
        'info': request.json['info']
    }
    tasks.append(task)
    return jsonify({'result': 'success', 'current tasks are': tasks})


@app.route('/get_task/', methods=['GET'])
@cross_origin(supports_credentials=True)
def get_task():
    if not request.args or 'id' not in request.args:
        return jsonify(tasks)
    else:
        task_id = request.args['id']
        task = list(filter(lambda t: t['id'] == int(task_id), tasks))
        return jsonify(task[0])


@app.route('/load_excel/', methods=['GET'])
@cross_origin(supports_credentials=True)
def load_excel():
    curPath = os.path.abspath('.')
    data = xlrd.open_workbook(os.path.join(curPath, "test.xlsx"))

    table = data.sheets()[0]
    rows = table.nrows

    content = []

    for i in range(rows):
        if i > 0:
            item_data = table.row_values(i)
            item_content = []
            for x in item_data:
                item_content.append(x)
            content.append(item_content)

    def createJson(content):
        jsonData = {}
        for item_data in content:
            jsonData[item_data[0]] = {}
            jsonData[item_data[0]]['name'] = item_data[0]
            jsonData[item_data[0]]['description'] = item_data[1]
            jsonData[item_data[0]]['field1'] = item_data[2]
            jsonData[item_data[0]]['field2'] = item_data[3]
        return jsonData

    def store(jsonStr):
        timeStrap = time.strftime("%a%b%d_%H_%M_%S_%Y", time.localtime());  # time strap
        name = "test_" + timeStrap + ".json";
        with open(name, 'w') as json_file:
            json.dump(jsonStr, json_file)

    jsonData = createJson(content)
    store(jsonData)
    jsonString = json.dumps(jsonData)
    print(jsonString)
    return jsonString


@app.route('/save_excel/', methods=['POST'])
@cross_origin(supports_credentials=True)
def save_excel():
    jsfile = "test"
    wb1 = Workbook()
    sheet = wb1.active
    data = json.load(open(jsfile + ".json", "r", encoding="utf-8"))
    values = list(data.values())
    for c, i in enumerate(values[0].keys()):
        sheet.cell(row=1, column=c + 1, value=i)
    for r, i in enumerate(values):
        row = r + 2
        for c, d in enumerate(values[r]):
            sheet.cell(row=row, column=c + 1, value=values[r].get(d))
    timeStrap = time.strftime("%a%b%d_%H_%M_%S_%Y", time.localtime());  # time strap
    fileOutName = "test_" + timeStrap + ".xlsx";
    wb1.save(fileOutName)
    status = 'api complete'
    print(status)
    return status


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8080, debug=True)